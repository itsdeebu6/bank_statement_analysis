import openpyxl as xl
from collections import OrderedDict
import datetime
from openpyxl.chart import (
    PieChart3D,
    Reference
)
def is_number(s):
   try:
      float(s) # for int, long and float
   except TypeError:
      return False
   except ValueError:
      try:
         complex(s) # for complex
      except ValueError:
         return False
   return True

def checkKey(dict, key):
   if key in dict.keys():
      return 1
   else:
      return 0

def catagories_the_expense (statement_filename_with_path):
   wb = xl.load_workbook(statement_filename_with_path)
   ws = wb[statement_sheetname]
   max_col = ws.max_column
   max_row = ws.max_row
   for i in range(1, max_row + 1):
      data_column = ws.cell(row = i, column = statement_date_column)
      data_column = str(data_column.value)
      #print(data_column)
      valid_transaction = 0
      category = ""
      try:
         datetime.datetime.strptime(data_column, statement_date_format)
         valid_transaction = 1
      except:
         print("skipping non transcation row")
      if valid_transaction == 1:
         cell_obj = ws.cell(row = i, column = statement_description_column)
         description = str(cell_obj.value)
         description = description.lower()
         for key, substring_list in expense_key_words.items():
            for substring in substring_list:
               substring = substring.lower()
               if substring in description:
                  category = key
                  break
            if category != "":
               break
         if category == "":
            category = "Other Transactions"
         category_to_update = ws.cell(row = i, column = max_col+1)
         category_to_update.value = category
         category = category.replace(" ", "_")
         if checkKey(expense_analyse_dict, category) == 0:
            expense_analyse_dict[category] = [0,0,0,0]
         
         #is credit or debit
         debit = 0
         credit = 0
         debit_column = ws.cell(row = i, column = statement_debit_column)
         debit_column = debit_column.value
         credit_column = ws.cell(row = i, column = statement_credit_column)
         credit_column = credit_column.value
         if is_number(debit_column) == True:
            debit = 1
            expense_analyse_dict[category][0] = expense_analyse_dict[category][0] + debit
            expense_analyse_dict[category][1] = expense_analyse_dict[category][1] + debit_column
         if is_number(credit_column) == True:
            credit = 1
            expense_analyse_dict[category][2] = expense_analyse_dict[category][2] + credit
            expense_analyse_dict[category][3] = expense_analyse_dict[category][3] + credit_column
   #print (expense_analyse_dict)     
         
   wb.save(statement_filename_with_path)
   wb.close()

def create_chart_for_expense_analysis (statement_filename_with_path, expense_analyse_dict):
   wb = xl.load_workbook(statement_filename_with_path)
   ws = wb.create_sheet("Expense_Analysis")
   ws.cell(row = 1, column = 1).value = "Expense Category"
   ws.cell(row = 1, column = 2).value = "Total debit transactions"
   ws.cell(row = 1, column = 3).value = "Total debit amount"
   ws.cell(row = 1, column = 4).value = "Total Credit transactions"
   ws.cell(row = 1, column = 5).value = "Total credit amount"
   i = 2
   j = 1
   total_keys = 0
   for key, expense_list in expense_analyse_dict.items():
      ws.cell(row = i, column = 1).value = key
      ws.cell(row = i, column = 2).value = expense_list[0]
      ws.cell(row = i, column = 3).value = expense_list[1]
      ws.cell(row = i, column = 4).value = expense_list[2]
      ws.cell(row = i, column = 5).value = expense_list[3]
      i+=1
      total_keys+=1

   chartposition = total_keys+2
   chartObj = ""
   chartObj = xl.chart.PieChart()
   labels = Reference(ws, min_col=1, min_row=2, max_row=total_keys+1)
   data = Reference(ws, min_col=2, min_row=1, max_row=total_keys+1)
   chartObj.add_data(data, titles_from_data=True)
   chartObj.set_categories(labels)
   chartObj.title = "Total Debits"
   ws.add_chart(chartObj, "A"+str(chartposition))

   chartObj = ""
   chartObj = xl.chart.PieChart()
   labels = Reference(ws, min_col=1, min_row=2, max_row=total_keys+1)
   data = Reference(ws, min_col=3, min_row=1, max_row=total_keys+1)
   chartObj.add_data(data, titles_from_data=True)
   chartObj.set_categories(labels)
   chartObj.title = "Debits Analysis"
   ws.add_chart(chartObj, "J"+str(chartposition))


   chartposition = chartposition + 20
   chartObj = ""
   chartObj = xl.chart.PieChart()
   labels = Reference(ws, min_col=1, min_row=2, max_row=total_keys+1)
   data = Reference(ws, min_col=4, min_row=1, max_row=total_keys+1)
   chartObj.add_data(data, titles_from_data=True)
   chartObj.set_categories(labels)
   chartObj.title = "Total Credits"
   ws.add_chart(chartObj, "A"+str(chartposition))

   chartObj = ""
   chartObj = xl.chart.PieChart()
   labels = Reference(ws, min_col=1, min_row=2, max_row=total_keys+1)
   data = Reference(ws, min_col=5, min_row=1, max_row=total_keys+1)
   chartObj.add_data(data, titles_from_data=True)
   chartObj.set_categories(labels)
   chartObj.title = "Credits Analysis"
   ws.add_chart(chartObj, "J"+str(chartposition))

   wb.save(statement_filename_with_path)
   wb.close()

   

################################ User Datas #############################

statement_date_column = 4
statement_date_format = r"%d/%m/%y"
statement_description_column = 2
statement_debit_column = 5
statement_credit_column = 6
statement_sheetname = "Sheet 1"

expense_key_words = OrderedDict()
expense_key_words['Salary_Income'] = ["SALARY", "ARRIS GROUP"]
expense_key_words['Food_payments'] = ["swiggy", "zomato", "ADYAR ANANDA", "AASAI DOSAI", "AMMAS", "ANJAPPAR"]
expense_key_words['Online_shopping'] = ["amazon", "flipkart", "nykaa", "HOPSCOTCH"]
expense_key_words['Cab_payments'] = ["ola", "uber"]
expense_key_words['Investment_savings'] = ["UTIMUTUAL", "RD INSTALLMENT", "UTIMF", "PRIYADASINI-SBIN"]
expense_key_words['Grocery_shop'] = ["GREENS FRESH", "SOBHA SUPER", "VINAYAKA HOT", "ESWARI STORES"]
expense_key_words['Online_grocery'] = ["SLURRP", "BIGBASKET", "DUNZO"]
expense_key_words['Fastag'] = ["FASTAG"]
expense_key_words['Cash Withdraw'] = ["NWD", "ATW"]
expense_key_words['To_Priya_Home_expense'] = ["PRIYADARSINI S", "DARSU3"]
expense_key_words['Medicals'] = ["MEDICA", "BENGALURU DRUG", "SRI MANJUNATHA MEDIC", "CHEMISTS", "1MG"]
expense_key_words['Fund_Transfer'] = ["NEFT", "imps", "FUNDS TRANSFER", "TPT-PAY", "PAYZAPP"]
expense_key_words['Spa_beautification'] = ["GREEN TRENDS", "LOGAMBAL", "LN ARTISTRY", "HAIR MASTERS"]
expense_key_words['Fuel'] = ["Ridhisu", "SHELL", "ANDAL SIR", "PETROL", "INDIANOIL", "BPCL", "THE SRINI"]
expense_key_words['Fixed_deposit'] = [" FD "]
expense_key_words['Travel'] = ["REDBUS"]
expense_key_words['Bike_Car_maintainance'] = ["ELITE MOTORS", "NEO", "MY TIRE STORE", "RAMANI CARS PRIV", "RISHI", "NAGARAJ"]
expense_key_words['Home_Rent'] = ["PRANAV MOHAN"]
expense_key_words['Kid_shopping'] = ["OH MY BABY", "firstcry", "FIRST CRY", "KIDZON"]
expense_key_words['Shopping'] = ["CENTRAL", "KHADIM", "POOJA FANCY", "THE CHENNAI", "VISHAL MEGA MART", "COLORS COLLECTION", "NOBLE FURNITURE", "SHOPPERS STOP", "RELIANCE TRENDS", "G K VALE", "THANGAMALI", "MENS AND WOMENS", "SREE RADHAA SILK", "ALUKKAS"]
expense_key_words['Utility_bills'] = ["VIDEOCON-DTH", "airtel", "VENKATESH", "DHAKSHAYINI", "ELECTRICITY", "JIO MOBILITY", "INDANEGAS", "NANDA FEEDS", "URBAN COMPANY", "URBANCLAP", "S R FRESH FISHES", "HOTSTAR"]
expense_key_words['Hospital'] = ["RAINBOW", "SAURABH", "NANO HOSPITAL", "DR SAHUS"]
expense_key_words['Credit_card_payment'] = ["5746", "CREDIT-CCPAY", "3005"]
expense_key_words['Other_Debit_card_payments'] = ["pos ", "6155"]
expense_key_words['Other_UPI_payments'] = ["upi"]


expense_analyse_dict = {}

################################ User Datas #############################


print('Enter your statement file path: ')
statement_filename_with_path = input()
catagories_the_expense (statement_filename_with_path)


print (expense_analyse_dict)
create_chart_for_expense_analysis (statement_filename_with_path, expense_analyse_dict)

